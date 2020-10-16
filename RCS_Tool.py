from openpyxl import load_workbook  #need to install
import pandas as pd   #need to install
from datetime import date
import sys

print('\
------------------------------------  INSTRUCTIONS ---------------------------------------------\n\
*******  THIS PROGRAM IS USED TO GENERATE RCS, PLEASE FOLLOW BELOW INSTRUCTIONS:         *******\n\
*******  1. TEMPLATE FILE NAME AS: RCS_template                                          *******\n\
*******  2. REMEDY FILE NAME AS: remedy                                                  *******\n\
*******  3. BTU FILE NAME AS: BTU                                                        *******\n\
*******  4. RND LTE FILE NAME AS: RND                                                    *******\n\
*******  5. RND UMTS FILE (IF APPLICABLE) NAME AS: RND_UMTS                              *******\n\
*******  6. ALL FILES IN SAME FOLDER AS THIS PYTHON PROGRAM                              *******\n\
*******  7. THIS PROGRAM WILL AUTO DETECT FILES                                          *******\n\
*******  Author: Ken Liang    email: ken1997528@hotmail.com    last update: 2019 Dec 19  *******\n\
----------------------------------  INSTRUCTIONS END  ------------------------------------------\n')
### file path
remedy_file = 'remedy.xlsx'
RND_file = 'RND.xls'
BTU_file = 'BTU.xlsm'
UMTS_file = 'RND_UMTS.xls'

### This section check if files are provided in the folder
while True:
    try:  ## check if UMTS file is provided
        umts_df1 = pd.read_excel(UMTS_file)
        UMTS = 1
    except:
        UMTS = 0
    try: ##try to load files into data frames
        remedy = pd.read_excel(remedy_file)
        rnd = pd.read_excel(RND_file)
        btu = pd.read_excel(BTU_file)
        if UMTS==1:
            print('AUTO FILE DETECTION: FILES (INCLUDING RND_UMTS) ARE IN PLACE! \n')
        else:
            print('AUTO FILE DETECTION: FILES ARE IN PLACE! (RND_UMTS NOT DETECTED)\n')
        break
    except:
        print('AUTO FILE DETECTION: FILE NOT FOUND!\nPLEASE READ ABOVE INSTRUCTIONS BEFORE PROCEED\n')
        input('if you have done the instructions, hit Enter to continue:')
        input('')
        print('\n')

### This section ask user for input
EMG = input('Please enter EMG (e.g T35JK) OR hit Enter to exit: ').upper()
if EMG=='':
    sys.exit()
Address = input('enter address (optional) OR hit Enter to skip: ')
Designer = input('enter your name as designer (optional) OR hit Enter to skip: ')

### this section check validity of EMG input by user, and validity of files provided in the folder (with 3 while loops)
while True: ##this while loop check if column EMG is in the remedy file and file contain info of site
    try:  ## load remedy file into data frame and try to look for column EMG
        remedy_col_label1 = 'EMG'
        remedy_col_label2 = 'Municipality_Name'
        remedy_col_label3 = 'PSAP'
        remedy = pd.read_excel(remedy_file) #load file into data frame
        remedy = remedy[remedy[remedy_col_label1]==EMG] #get rows containing input EMG, if file contains no column EMG, exception thrown
        if remedy.empty: #if EMG is not in the EMG column, prompt user to put in the correct file
            print('\nERROR: ***EMG NOT FOUND IN PROVIDED REMEDY FILE*** \n')
            print('if you enter the wrong EMG, enter 1 to re-enter EMG\n')
            print('if the file does not have info of the site, simply replace with the correct file (make sure to change file name) and hit Enter: ')
            prompt = input('')
            if prompt=='1':
                EMG = input('Please enter EMG (e.g T35JK): ').upper()
        else: # file has the info of the site, now get data needed: Municipality, PSAP
            remedy.reset_index(inplace=True)
            ###look for Municipality and PSAP
            for i in [remedy_col_label2, remedy_col_label3]:
                temp = remedy[i]
                if i == remedy_col_label2:
                    Municipality = temp[0]
                elif i == remedy_col_label3:
                    PSAP = temp[0]
            break
    except:
        print('\nERROR: ***REMEDY FILE MISSING COLUMN: EMG***\n')
        print('Instruction 1: make sure the remedy file has a column that is containing EMG info\n')
        print('Instruction 2: make sure the EMG column label is named EMG\n')
        prompt = input('hit Enter to continue if the above instructions have been achieved: ')
        print('\n')

while True:### check if RND file containing input EMG
    try:
        sheet_PCI = 'PCI'
        sheet_eUtran = 'eUtran Parameters'
        PCI_col_label1 = 'EutranCellFDDId'
        PCI_col_label2 = 'PCI'
        eUtran_col_label1 = 'EutranCellFDDId'
        eUtran_col_label2 = 'beamDirection'
        rnd1 = pd.read_excel(RND_file, sheet_name=sheet_PCI)  ##try to load worksheet PCI, if has no sheet PCI, exception thrown
        rnd2 = pd.read_excel(RND_file, sheet_name=sheet_eUtran)  ##try to load sheet eUtran Parameters, if has no sheet PCI, exception thrown
        ##try to look for column EutranCellFDDId from sheet PCI (if does not have this column, exception thrown) and check if it contain EMG of site
        rnd1 = rnd1[rnd1[PCI_col_label1].str.contains(pat='^' + EMG, regex=True)]
        ##try to look for column EutranCellFDDId from sheet eUtran Parameters (if does not have this column, exception thrown) and check if it contain EMG of site
        rnd2 = rnd2[rnd2[eUtran_col_label1].str.contains(pat='^' + EMG, regex=True)]
        if rnd1.empty or rnd2.empty: ##this file does not contain info of the site (could not find the input EMG from the column EMG)
            print('\nERROR: ***THIS SITE IS NOT IN THE PROVIDED RND FILE*** \n')
            print('please provide the correct RND file')
            prompt = input('hit Enter to continue if the above instruction has been achieved: ')
        else: ##this file has info of the site, get the data needed: PCI, Azimuth
            # get PCI
            PCI_dict = {}
            for i in range(len(rnd1)):
                PCI_dict.update({rnd1[PCI_col_label1][i][-1]: rnd1[PCI_col_label2][i]})
            # get Azimuth
            Az_dict = {}
            for i in range(len(rnd2)):
                Az_dict.update({rnd2[eUtran_col_label1][i][-1]: rnd2[eUtran_col_label2][i]})
            break
    except: ## exception to be thrown
        print('\nERROR: ***RND FILE MISSING SHEET(S): PCI AND/OR eUtran Parameters***\n')
        print('Instruction 1: make sure the RND file has worksheets named PCI and eUtran Parameters\n')
        print('Instruction 2: make sure the worksheet PCI has a column PCI and a column EutranCellFDDId\n')
        print('Instruction 3: make sure the worksheet eUtran Parameters has a column beamDirection and a column EutranCellFDDId\n')
        prompt = input('hit Enter to continue if the above instructions have been achieved: ')
        print('\n')

while True:
    try:##load BTU, try to look for the columns, if does not have, exception thrown
        col_label1 = 'LTE CELL Name\n(CELL_SITE)'
        col_label2 = 'EUTRAN_CELL_ID'
        col_label3 = 'EUTRAN_CELL_ID\n(HEX value)'
        col_label4 = 'ESRD'
        col_label5 = 'ESN'
        col_label6 = 'Site Name'
        col_label7 = 'EMG'
        col_label8 = 'Location Code'
        row_number = 12
        btu = pd.read_excel(BTU_file, header=row_number-1, usecols=[col_label1, col_label2, col_label3, col_label4, col_label5, col_label6, col_label7, col_label8])
        btu = btu.loc[btu[col_label7] == EMG]  ##select data of the site
        if btu.empty:  ## this BTU file does not have info of the site
            print('\n ERROR: ***THIS SITE IS NOT IN THE PROVIDED BTU FILE*** \n')
            print('please provide the correct BTU file')
            prompt = input('hit Enter to continue if the above instruction has been achieved: ')
        else: #file has the info of the site, get data needed: 'Cell Name', 'Sector', 'Eutran Cell ID', 'Eutran Cell ID (HEX)', 'Frequency','EARFCN DL', 'PCI', 'Azimuth', 'MIMO', 'ESRD', 'ESN']
            freq_dict = {
                '1': '700',
                '2': '700',
                '3': '700',
                '4': '2100',
                '5': '2100',
                '6': '2100',
                '7': '2600',
                '8': '2600',
                '9': '2600',
                'A': '700',
                'B': '700',
                'C': '700',
                'D': '1900',
                'E': '1900',
                'F': '1900',
                'G': '850',
                'H': '850',
                'J': '850',
                'K': '1900',
                'L': '1900',
                'M': '1900',
                'N': '2100',
                'P': '2100',
                'Q': '2100',
            }; EARFCN_dict = {
                '1': '5765',
                '2': '5765',
                '3': '5765',
                '4': '2025',
                '5': '2025',
                '6': '2025',
                '7': '3050',
                '8': '3050',
                '9': '3050',
                'A': '5060',
                'B': '5060',
                'C': '5060',
                'D': '675',
                'E': '675',
                'F': '675',
                'G': '2435',
                'H': '2435',
                'J': '2435',
                'K': '1075',
                'L': '1075',
                'M': '1075',
                'N': '2225',
                'P': '2225',
                'Q': '2225',
            }; Sector_dict = {
                '1': 'main1',
                '2': 'main2',
                '3': 'main3',
                '4': 'main1',
                '5': 'main2',
                '6': 'main3',
                '7': 'main1',
                '8': 'main2',
                '9': 'main3',
                'A': 'main1',
                'B': 'main2',
                'C': 'main3',
                'D': 'offset1',
                'E': 'offset2',
                'F': 'offset3',
                'G': 'offset1',
                'H': 'offset2',
                'J': 'offset3',
                'K': 'offset1',
                'L': 'offset2',
                'M': 'offset3',
                'N': 'main1',
                'P': 'main2',
                'Q': 'main3',
            }; MIMO_dict = {
                '1': '4x4',
                '2': '4x4',
                '3': '4x4',
                '4': '4x4',
                '5': '4x4',
                '6': '4x4',
                '7': '4x4',
                '8': '4x4',
                '9': '4x4',
                'A': '4x4',
                'B': '4x4',
                'C': '4x4',
                'D': '4x4',
                'E': '4x4',
                'F': '4x4',
                'G': '2x4',
                'H': '2x4',
                'J': '2x4',
                'K': '4x0',
                'L': '4x0',
                'M': '4x0',
                'N': '4x0',
                'P': '4x0',
                'Q': '4x0',
            }  # mapping dict for frequency, EARFCN, Sector number, MIMO
            Freq_col = []; EARFCN_col = []; Sector_col = []; PCI_col = []; Az_col = []; MIMO_col = []  # empty lists to store data later
            btu.reset_index(drop=True, inplace=True)  ##reset index of the dataframe
            LocationCode = btu[col_label8][0]  ##get location code of the site, e.g C0001
            SiteName = btu[col_label6][0]  ## get site name
            btu.drop(columns=col_label8, inplace=True)  # no need Location Code column
            btu.drop(columns=col_label6, inplace=True)  # no need Site Name column
            btu.drop(columns=col_label7, inplace=True)  # no need EMG column
            ### this loop get the last character from Cell Name as key, then create Freq, EARFCN, Sector, PCI, Az columns base on dictionaries
            for i in range(len(btu)):
                key = btu[col_label1][i][-1]
                Freq_col.append(freq_dict.get(key))
                EARFCN_col.append(EARFCN_dict.get(key))
                Sector_col.append(Sector_dict.get(key))
                PCI_col.append(PCI_dict.get(key))
                Az_col.append(Az_dict.get(key))
                MIMO_col.append(MIMO_dict.get(key))
            btu.insert(loc=1, column='Sector', value=Sector_col)   # insert Sector column in dataframe
            btu.insert(loc=4, column='Frequency', value=Freq_col)  # insert Frequency column in dataframe
            btu.insert(loc=5, column='EARFCN DL', value=EARFCN_col)  # insert EARFCN DL column in dataframe
            btu.insert(loc=6, column='PCI', value=PCI_col)    # insert PCI column in dataframe
            btu.insert(loc=7, column='Azimuth', value=Az_col)  # insert Azimuth column in dataframe
            btu.insert(loc=8, column='MIMO', value=MIMO_col)  # insert MIMO column in dataframe
            ##rename column labels
            btu.columns = ['Cell Name', 'Sector', 'Eutran Cell ID', 'Eutran Cell ID (HEX)', 'Frequency',
                           'EARFCN DL', 'PCI', 'Azimuth', 'MIMO', 'ESRD', 'ESN']
            btu.ESN = btu.ESN.astype(int)  ##change type from float to int
            btu.Frequency = btu.Frequency.astype(int)  ##change type from float to int
            btu['EARFCN DL'] = btu['EARFCN DL'].astype(int)  ##change type from float to int
            NumbRows = btu.shape[0]  # get number of rows in this dataframe for later excel editing
            NumbCols = btu.shape[1]  # get number of rows in this dataframe for later excel editing
            break
    except:
        print('\nERROR: ***BTU FILE MISSING COLUMNS***\n')
        print('\
Instruction: BTU file should contain the following columns: \
1. \'LTE CELL Name\\n(CELL_SITE)\'    2. \'EUTRAN_CELL_ID\'   3. \'EUTRAN_CELL_ID\\n(HEX value)\'\n\
4. \'ESRD\'    5. \'ESN\'    6. \'Site Name\'    7. \'EMG\'    8. \'Location Code\'')
        prompt = input('hit Enter to continue if the above instruction have been achieved: ')
        print('\n')

### this section relates to UMTS file (if provided)
if UMTS==1:
    umts_sheet1 = 'EutranFreqRelation'
    umts_sheet2 = 'dyn RN'
    umts_sheet3 = 'Lac-Sac-Rac'
    Eutran_sheet_col1 = 'Utrancell'
    Eutran_sheet_col2 = 'RNC'
    dyn_RN_sheet_col1 = 'sector Number'
    dyn_RN_sheet_col2 = 'cell Identity'
    dyn_RN_sheet_col3 = 'uarfcn Dl'
    dyn_RN_sheet_col4 = 'scrambling code'
    dyn_RN_sheet_col5 = 'beam Direction'
    Lac_Sac_Rac_sheet_col1 = 'UMTS LAC'
    umts_df1 = pd.read_excel(UMTS_file, sheet_name=umts_sheet1)  ###load sheet 'EutranFreqRelation' in data frame
    output_df = umts_df1[[Eutran_sheet_col1, Eutran_sheet_col2]].copy(deep=True)  ##extract columns needed and store copy in output_df
    umts_df2 = pd.read_excel(UMTS_file, sheet_name=umts_sheet2)    ###load sheet 'dyn RN' in data frame
    umts_df3 = pd.read_excel(UMTS_file, sheet_name=umts_sheet3)  ### load sheet
    search_col = ['sector Number', 'cell Identity', 'scrambling code', 'uarfcn Dl', 'beam Direction']
    output_df.insert(loc=1, column='Sector', value=umts_df2[dyn_RN_sheet_col1]) #insert Sector column
    ### CGI = 302-720-<umts lac>-<cellID>
    output_df.insert(loc=3, column='CGI', value='302-720-' + umts_df3[Lac_Sac_Rac_sheet_col1].astype(str) + '-' + umts_df2[dyn_RN_sheet_col2].astype(str)) #insert CGI
    ### temperarily use value of Utrancell for Frequency, will edit later
    output_df.insert(loc=4, column='Frequency', value=umts_df1[Eutran_sheet_col1])  # insert dummuy Frequency col for later to edit
    for i in range(len(output_df)):
        if output_df['Frequency'][i][-1] == '1' or output_df['Frequency'][i][-1] == '2' or output_df['Frequency'][i][-1] == '3':
            output_df.loc[i,['Frequency']] = 'HSPA850'
        else:
            output_df.loc[i,['Frequency']] = 'HSPA1900'
    output_df.insert(loc=5, column='UARFCN DL', value=umts_df2[dyn_RN_sheet_col3])  # insert UARFCN DL
    output_df.insert(loc=6, column='SC', value=umts_df2[dyn_RN_sheet_col4])   # insert SC
    output_df.insert(loc=7, column='Azimuth', value=umts_df2[dyn_RN_sheet_col5])   # insert Azimuth
    ### temperarily use value of Utrancell for ESRD and ESN, will edit later
    output_df.insert(loc=8, column='ESRD', value=umts_df1[Eutran_sheet_col1])
    ### temperarily use value of Utrancell for Frequency, will edit later
    output_df.insert(loc=9, column='ESN', value=umts_df1[Eutran_sheet_col1])
    ### look for ESRD  and ESN base on Azimuth
    Az = []
    for i in range(len(output_df)):
        Az = output_df['Azimuth'][i]
        temp = btu.loc[btu.Azimuth==Az].copy(deep=True)
        temp.reset_index(drop=True, inplace=True)
        output_df.loc[i,['ESRD']] = temp['ESRD'][i]
        output_df.loc[i, ['ESN']] = temp['ESN'][i]
    NumbRows_UMTS = output_df.shape[0]
    NumbCols_UMTS = output_df.shape[1]
    #pd.set_option('display.max_columns', 10)
    #print(output_df)
#['Cell Name', 'Sector', 'RNC', 'CGI', 'Frequency', 'UARFCN DL', 'SC', 'Azimuth', 'ESRD', 'ESN']




###inport data to excel
template_path = 'RCS_template.xlsx'
output_path = LocationCode + '_' + EMG + '_RCS_RevA-' + date.today().strftime("%Y-%m-%d") +'.xlsx'

writer = pd.ExcelWriter(output_path, engine='openpyxl')  #create output file
writer.book = load_workbook(template_path)  #load template
btu.to_excel(writer, sheet_name='RCS', index=False)
if UMTS==1:
    output_df.to_excel(writer, sheet_name='RCS_UMTS', index=False)
writer.save()

###fill out first sheet
workbook = load_workbook(output_path)
#editing first worksheet
worksheet1 = workbook['Site']
worksheet1['B3'] = date.today().strftime("%d-%b")
worksheet1['B4'] = Designer
#worksheet1['B5'] = proj numb
worksheet1['B6'] = SiteName
worksheet1['B7'] = LocationCode
worksheet1['B8'] = EMG
worksheet1['B9'] = Address
worksheet1['B10'] = PSAP
worksheet1['B11'] = Municipality
# editing LTE worksheet, moving data from RCS sheet to LTE sheet
worksheetLTE = workbook['LTE']
worksheetRCS = workbook['RCS']
for i in range(NumbCols+1):
    for j in range(NumbRows+1):
        worksheetLTE.cell(row=j+1, column=i+1, value=worksheetRCS.cell(row=j + 1, column=i + 1).value)

# editing UMTS worksheet, moving data from RCS_UMTS sheet to UMTS sheet
if UMTS==1:
    worksheetUMTS = workbook['UMTS']
    worksheetRCS_UMTS = workbook['RCS_UMTS']
    for i in range(NumbCols_UMTS + 1):
        for j in range(NumbRows_UMTS + 1):
            worksheetUMTS.cell(row=j + 1, column=i + 1, value=worksheetRCS_UMTS.cell(row=j + 1, column=i + 1).value)
    workbook.remove(worksheetRCS_UMTS)
else:
    workbook.remove(workbook['UMTS'])
workbook.remove(worksheetRCS)
workbook.save(output_path)


print('\n----------  RCS FILE HAS BEEN GENERATED, REMEMBER TO ATTACH MAP  ----------')


